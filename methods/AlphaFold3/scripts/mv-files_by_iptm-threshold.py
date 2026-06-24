#!/usr/bin/env python3

import sys
import shutil
from pathlib import Path
from openpyxl import load_workbook

xlsx_file, src_dir, dst_dir, threshold = sys.argv[1:]
threshold = float(threshold)

src_dir = Path(src_dir)
dst_dir = Path(dst_dir)
dst_dir.mkdir(exist_ok=True)

wb = load_workbook(xlsx_file)
ws = wb.active

# read header manually (old openpyxl returns cell objects)
header = [c.value for c in ws[1]]

name_col = header.index("Folder Name")
iptm_col = header.index("iPTM")

copied = 0

# iterate rows manually (no values_only)
for row in ws.iter_rows(min_row=2):
    folder_name = row[name_col].value
    iptm = row[iptm_col].value

    if folder_name is None or iptm is None:
        continue

    try:
        if float(iptm) < threshold:
            for f in src_dir.glob(str(folder_name) + "*"):
                if f.is_file():
                    shutil.copy2(f, dst_dir / f.name)
                    print("Copied:", f.name)
                    copied += 1
    except ValueError:
        continue

print(f"\nDone. Total files copied: {copied}")