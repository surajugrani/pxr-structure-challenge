#!/usr/bin/env python3

import os
import sys
import shutil
import openpyxl


def read_results(folder):
    results_file = os.path.join(folder, "results.xlsx")

    if not os.path.exists(results_file):
        raise IOError("Missing results file: %s" % results_file)

    wb = openpyxl.load_workbook(results_file)
    ws = wb.active

    # ---- Get header row safely (old openpyxl compatible) ----
    header_row = None
    for row in ws.rows:
        header_row = row
        break

    if header_row is None:
        raise RuntimeError("Empty worksheet: %s" % results_file)

    headers = []
    for cell in header_row:
        headers.append(cell.value)

    try:
        folder_col = headers.index("Folder Name")
        iptm_col = headers.index("iPTM")
    except ValueError:
        raise RuntimeError(
            "%s must contain columns 'Folder Name' and 'iPTM'"
            % results_file
        )

    results = {}

    for row in ws.iter_rows(min_row=2):

        folder_value = row[folder_col].value
        iptm_value = row[iptm_col].value

        if folder_value is None:
            continue

        complex_name = str(folder_value).strip()
        iptm = float(iptm_value)

        results[complex_name] = iptm

    return results


def main():

    if len(sys.argv) < 4:
        print(
            "Usage: python pick-best-iptm.py "
            "<folder1> <folder2> ... <output_dir>"
        )
        sys.exit(1)

    input_folders = sys.argv[1:-1]
    output_dir = sys.argv[-1]

    os.makedirs(output_dir, exist_ok=True)

    best_models = {}

    for folder_index, folder in enumerate(input_folders):

        print("Reading:", folder)

        results = read_results(folder)

        folder_path = os.path.abspath(folder)
        folder_name = os.path.basename(folder_path)

        for complex_name, iptm in results.items():

            if complex_name not in best_models:
                best_models[complex_name] = {
                    "iptm": iptm,
                    "folder_path": folder_path,
                    "folder_name": folder_name,
                    "index": folder_index,
                }

            elif iptm > best_models[complex_name]["iptm"]:
                best_models[complex_name] = {
                    "iptm": iptm,
                    "folder_path": folder_path,
                    "folder_name": folder_name,
                    "index": folder_index,
                }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Best Models"

    ws.append(["Complex", "Source Folder", "iPTM"])

    copied = 0

    for complex_name in sorted(best_models.keys()):

        info = best_models[complex_name]

        src = os.path.join(info["folder_path"], complex_name + ".pdb")
        dst = os.path.join(output_dir, complex_name + ".pdb")

        if not os.path.exists(src):
            raise IOError("Missing PDB: %s" % src)

        shutil.copy2(src, dst)
        copied += 1

        ws.append([
            complex_name,
            info["folder_name"],
            info["iptm"],
        ])

        print(
            "Copied %s from %s (iPTM=%.3f)"
            % (complex_name, info["folder_name"], info["iptm"])
        )

    out_xlsx = os.path.join(output_dir, "best_models.xlsx")
    wb.save(out_xlsx)

    print("\nCopied %d PDB files" % copied)
    print("Wrote:", out_xlsx)


if __name__ == "__main__":
    main()

##!/usr/bin/env python3
#
#import os
#import sys
#import shutil
#import openpyxl
#
#
#def read_results(folder):
#    results_file = os.path.join(folder, "results.xlsx")
#
#    if not os.path.exists(results_file):
#        raise IOError("Missing results file: %s" % results_file)
#
#    wb = openpyxl.load_workbook(results_file)
#    ws = wb.active
#
#    # ---- FIX: safely get header row (no ws.rows[0]) ----
#    header_row = None
#    for i, row in enumerate(ws.rows):
#        header_row = row
#        break
#
#    if header_row is None:
#        raise RuntimeError("Empty worksheet: %s" % results_file)
#
#    headers = []
#    for cell in header_row:
#        headers.append(cell.value)
#
#    try:
#        folder_col = headers.index("Folder Name")
#        iptm_col = headers.index("iPTM")
#    except ValueError:
#        raise RuntimeError(
#            "%s must contain columns 'Folder Name' and 'iPTM'"
#            % results_file
#        )
#
#    results = {}
#
#    # ---- data rows ----
#    for row in ws.iter_rows(min_row=2):
#
#        folder_value = row[folder_col].value
#        iptm_value = row[iptm_col].value
#
#        if folder_value is None:
#            continue
#
#        complex_name = str(folder_value).strip()
#        iptm = float(iptm_value)
#
#        results[complex_name] = iptm
#
#    return results
#
#
#def main():
#
#    if len(sys.argv) < 4:
#        print(
#            "Usage: python pick-best-iptm.py "
#            "<folder1> <folder2> ... <output_dir>"
#        )
#        sys.exit(1)
#
#    input_folders = sys.argv[1:-1]
#    output_dir = sys.argv[-1]
#
#    os.makedirs(output_dir, exist_ok=True)
#
#    best_models = {}
#
#    for folder_index, folder in enumerate(input_folders):
#
#        print("Reading:", folder)
#
#        results = read_results(folder)
#
#        for complex_name, iptm in results.items():
#
#            if complex_name not in best_models:
#                best_models[complex_name] = {
#                    "iptm": iptm,
#                    "folder_path": os.path.abspath(folder),
#                    "folder_name": os.path.basename(os.path.abspath(folder)),
#                    "index": folder_index,
#                }
#
#            elif iptm > best_models[complex_name]["iptm"]:
#                best_models[complex_name] = {
#                    "iptm": iptm,
#                    "folder_path": os.path.abspath(folder),
#                    "folder_name": os.path.basename(os.path.abspath(folder)),
#                    "index": folder_index,
#                }
#
#    wb = openpyxl.Workbook()
#    ws = wb.active
#    ws.title = "Best Models"
#
#    ws.append(["Complex", "Source Folder", "iPTM"])
#
#    copied = 0
#
#    for complex_name in sorted(best_models):
#
#        info = best_models[complex_name]
#
#        src = os.path.join(info["folder"], complex_name + ".pdb")
#        dst = os.path.join(output_dir, complex_name + ".pdb")
#
#        if not os.path.exists(src):
#            raise IOError("Missing PDB: %s" % src)
#
#        shutil.copy2(src, dst)
#        copied += 1
#
#        ws.append([
#            complex_name,
#            os.path.basename(info["folder"]),
#            info["iptm"],
#        ])
#
#        print(
#            "Copied %s from %s (iPTM=%.3f)"
#            % (complex_name, info["folder"], info["iptm"])
#        )
#
#    out_xlsx = os.path.join(output_dir, "best_models.xlsx")
#    wb.save(out_xlsx)
#
#    print("\nCopied %d PDB files" % copied)
#    print("Wrote:", out_xlsx)
#
#
#if __name__ == "__main__":
#    main()